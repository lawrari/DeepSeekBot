from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from md2tgmd import escape


class Formattor:

    @staticmethod
    def format_text(text):
        text = escape(text)

        text = text.replace(" •", " –")

        # Check table
        lines = text.split("\n")
        i = 0

        files = []

        while i < len(lines):
            if (
                ((lines[i].startswith("\\|") and lines[i].endswith("\\|")) or (lines[i].startswith("\\+") and lines[i].endswith("\\+")) or (lines[i].startswith("|") and lines[i].endswith("|")) or (lines[i].startswith("+") and lines[i].endswith("+")) )
                and
                i + 2 < len(lines) and 
                ((lines[i+1].startswith("\\|") and lines[i+1].endswith("\\|")) or (lines[i+1].startswith("\\+") and lines[i+1].endswith("\\+")) or (lines[i+1].startswith("|") and lines[i+1].endswith("|")) or (lines[i+1].startswith("+") and lines[i+1].endswith("+")) )
                and
                ( (lines[i+2].startswith("\\|") and lines[i+2].endswith("\\|")) or (lines[i+2].startswith("\\+") and lines[i+2].endswith("\\+")) or (lines[i+2].startswith("|") and lines[i+2].endswith("|")) or (lines[i+2].startswith("+") and lines[i+2].endswith("+")) )):
                j = i + 3

                table = []

                while j < len(lines) and  ( (lines[j].startswith("\\|") and lines[j].endswith("\\|")) or (lines[j].startswith("\\+") and lines[j].endswith("\\+")) or (lines[j].startswith("|") and lines[j].endswith("|")) or (lines[j].startswith("+") and lines[j].endswith("+")) ):
                    j += 1
                
                for k in range(i, j):
                    table.append(lines[k].replace("\\", "").replace("+", "|"))

                lines = lines[:i] + ["Таблица " + f"table\\_{len(files)}\\.xlsx"] + lines[j:]

                a = map(lambda x: list(map(lambda y: y.strip(), x.split("|")[1:-1])), table)

                wb = Workbook()
                ws = wb.active

                for row in list(a):
                    ws.append(row)

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column
                    column_letter = get_column_letter(column)
                    
                    for cell in col:
                        try:
                            # Пытаемся получить длину значения ячейки
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                byte_data = BytesIO()

                wb.save(byte_data)

                excel_bytes = byte_data.getvalue()

                filename = f"table_{len(files)}.xlsx"
                data = excel_bytes
                
                files.append(( filename, data ))
            else:
                i += 1

        res = "\n".join(lines)

        res1 = ""

        i = 0
        while i < len(res):
            if res[i] == "`":
                j = i + 1
                while j < len(res) and res[j] != "`":
                    j += 1
                if j < len(res):
                    a = str(res[i+1:j]) \
                    .replace(".", "\\.") \
                    .replace("~", "\\~")

                    res1 += "`" + a + "`"
                    i = j
            else:
                res1 += res[i]
            i += 1
        
        return ( res1, files )

